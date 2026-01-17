# core/etsy_exporter.py

from __future__ import annotations

import csv
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

from core.product_schema import ProductDraft


def product_to_etsy_row(product: ProductDraft, image_files: List[Dict[str, Any]] = None) -> Dict[str, Any]:
    """
    将 ProductDraft 转换为 Etsy 填写表格的一行数据。
    
    Args:
        product: ProductDraft 对象
        image_files: 图片文件列表（从 Drive 获取的文件信息），用于生成图片文件名
    
    Returns:
        包含所有 Etsy 字段的字典
    """
    # 处理图片：从 image_files 或 product.image_paths 获取
    image_names = []
    if image_files:
        # 使用 Drive 文件列表中的文件名
        for img_file in image_files[:10]:  # 最多 10 张图片
            image_names.append(img_file.get("name", ""))
    elif product.image_paths:
        # 使用本地路径中的文件名
        for img_path in product.image_paths[:10]:
            path_obj = Path(img_path)
            image_names.append(path_obj.name if path_obj.name else img_path)
    
    # 确保至少有 10 个图片列（用空字符串填充）
    while len(image_names) < 10:
        image_names.append("")
    
    # 构建行数据
    row = {
        "title": product.title,
        "description": product.description,
        "price": str(product.price),
        "currency": product.currency,
        "quantity": str(product.quantity),
        "category": product.category or product.product_type,
        "taxonomy_id": str(product.taxonomy_id) if product.taxonomy_id else "",
        "materials": ", ".join(product.materials) if product.materials else "",
        "tags": ", ".join(product.tags) if product.tags else "",
        "who_made": product.who_made,
        "when_made": product.when_made,
        "shop_section": "",  # 需要手动填写
        "shipping_profile": "",  # 需要手动填写
        "SKU": product.id,  # 使用 folder_id 作为 SKU
        "processing_time": "",  # 需要手动填写（天数）
    }
    
    # 添加图片列
    for i in range(10):
        row[f"image_{i+1}"] = image_names[i] if i < len(image_names) else ""
    
    return row


def export_products_to_csv(
    products: List[tuple[ProductDraft, List[Dict[str, Any]]]],
    output_path: Path
) -> Path:
    """
    将多个 ProductDraft 导出为 CSV 文件。
    
    Args:
        products: ProductDraft 和对应图片文件列表的元组列表
        output_path: 输出 CSV 文件路径
    
    Returns:
        输出文件路径
    """
    if not products:
        raise ValueError("没有产品需要导出")
    
    # 定义列顺序
    columns = [
        "title",
        "description",
        "price",
        "currency",
        "quantity",
        "category",
        "taxonomy_id",
        "materials",
        "tags",
        "image_1",
        "image_2",
        "image_3",
        "image_4",
        "image_5",
        "image_6",
        "image_7",
        "image_8",
        "image_9",
        "image_10",
        "who_made",
        "when_made",
        "shop_section",
        "shipping_profile",
        "SKU",
        "processing_time",
    ]
    
    # 确保输出目录存在
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # 写入 CSV
    with output_path.open("w", newline="", encoding="utf-8-sig") as f:  # utf-8-sig 支持 Excel 中文
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        
        for product, image_files in products:
            row = product_to_etsy_row(product, image_files)
            writer.writerow(row)
    
    return output_path


def export_products_to_excel(
    products: List[tuple[ProductDraft, List[Dict[str, Any]]]],
    output_path: Path
) -> Path:
    """
    将多个 ProductDraft 导出为 Excel 文件（需要 openpyxl）。
    
    Args:
        products: ProductDraft 和对应图片文件列表的元组列表
        output_path: 输出 Excel 文件路径
    
    Returns:
        输出文件路径
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        raise ImportError(
            "需要安装 openpyxl 库才能导出 Excel 文件。\n"
            "运行: pip install openpyxl"
        )
    
    if not products:
        raise ValueError("没有产品需要导出")
    
    # 定义列顺序和列宽
    columns = [
        ("title", "标题", 40),
        ("description", "详细描述", 80),
        ("price", "价格", 12),
        ("currency", "币种", 10),
        ("quantity", "库存", 10),
        ("category", "类别", 20),
        ("taxonomy_id", "分类编号", 15),
        ("materials", "材料", 40),
        ("tags", "标签", 50),
        ("image_1", "图片1", 30),
        ("image_2", "图片2", 30),
        ("image_3", "图片3", 30),
        ("image_4", "图片4", 30),
        ("image_5", "图片5", 30),
        ("image_6", "图片6", 30),
        ("image_7", "图片7", 30),
        ("image_8", "图片8", 30),
        ("image_9", "图片9", 30),
        ("image_10", "图片10", 30),
        ("who_made", "谁制作", 15),
        ("when_made", "何时制作", 20),
        ("shop_section", "店铺分区", 20),
        ("shipping_profile", "运费模板", 25),
        ("SKU", "SKU", 30),
        ("processing_time", "处理时间", 15),
    ]
    
    # 创建 workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Etsy 商品列表"
    
    # 写入表头（中英文）
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col_idx, (field_name, chinese_name, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = f"{chinese_name}\n({field_name})"
        cell.font = header_font
        cell.alignment = header_alignment
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    # 写入数据
    for row_idx, (product, image_files) in enumerate(products, start=2):
        row_data = product_to_etsy_row(product, image_files)
        
        for col_idx, (field_name, _, _) in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = row_data.get(field_name, "")
            cell.value = value
            # 描述列自动换行
            if field_name == "description":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    # 确保输出目录存在
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # 保存文件
    wb.save(output_path)
    
    return output_path



